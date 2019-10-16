from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Alignment, Font
from django.http.response import HttpResponse

from .Timemanager import timeCompare


class DownloadExcel:
    def __init__(self):
        pass
    def download_excel(self, employee_list):
        bold = Font(bold=True)
        output = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=utf-8;')
        file_name = "Employee Details.xlsx"
        output['Content-Disposition'] = 'attachment; filename=' + file_name
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 5
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 5
        ws.column_dimensions['M'].width = 10
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 10
        ws.column_dimensions['Q'].width = 10
        row = 1
        ws.merge_cells('A{0}:E{0}'.format(str(row)))
        ws['A'+str(row)].font = bold
        ws['A'+str(row)].value = 'EMPLOYEE TIMELINE'
        ws['A'+str(row)].alignment = Alignment(horizontal='center')

        ws.merge_cells('G{0}:K{0}'.format(str(row)))
        ws['G'+str(row)].font = bold
        ws['G'+str(row)].value = 'EMPLOYEE TIMELINE'
        ws['G'+str(row)].alignment = Alignment(horizontal='center')

        ws.merge_cells('M{0}:P{0}'.format(str(row)))
        ws['M'+str(row)].font = bold
        ws['M'+str(row)].value = 'EMPLOYEE TIMELINE'
        ws['M'+str(row)].alignment = Alignment(horizontal='center')

        row_green =row_yellow= row_red = 3
        row+=2
        ws['A'+str(row)].font = bold
        ws['A'+str(row)].value = "Id"
        ws['B'+str(row)].font = bold
        ws['B'+str(row)].value = "Name"
        ws['B'+str(row)].alignment = Alignment(wrapText=True)
        ws['C'+str(row)].font = bold
        ws['C'+str(row)].value = "Email"
        ws['C'+str(row)].alignment = Alignment(wrapText=True)
        ws['D'+str(row)].font = bold
        ws['D'+str(row)].value = "Time from"
        ws['E'+str(row)].font = bold
        ws['E'+str(row)].value = "Time to"
        ws['G'+str(row)].font = bold
        ws['G'+str(row)].value = "Id"
        ws['H'+str(row)].font = bold
        ws['H'+str(row)].value = "Name"
        ws['H'+str(row)].alignment = Alignment(wrapText=True)
        ws['I'+str(row)].font = bold
        ws['I'+str(row)].value = "Email"
        ws['I'+str(row)].alignment = Alignment(wrapText=True)
        ws['J'+str(row)].font = bold
        ws['J'+str(row)].value = "Time from"
        ws['K'+str(row)].font = bold
        ws['K'+str(row)].value = "Time to"
        ws['M'+str(row)].font = bold
        ws['M'+str(row)].value = "Id"
        ws['N'+str(row)].font = bold
        ws['N'+str(row)].value = "Name"
        ws['N'+str(row)].alignment = Alignment(wrapText=True)
        ws['O'+str(row)].font = bold
        ws['O'+str(row)].value = "Email"
        ws['O'+str(row)].alignment = Alignment(wrapText=True)
        ws['P'+str(row)].font = bold
        ws['P'+str(row)].value = "Time from"
        ws['Q'+str(row)].font = bold
        ws['Q'+str(row)].value = "Time to"

        redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
        greenFill = PatternFill(start_color='78f671',
                   end_color='78f671',
                   fill_type='solid')
        yellowFill = PatternFill(start_color='f8ed05',
                   end_color='f8ed05',
                   fill_type='solid')

        if employee_list:
            for key, value in employee_list.items():
                dto = EmployeeDetailsDTO()
                dto.empid = key
                dto.name = value['name']
                dto.email = value['email']
                dto.time_from = value['time_from']
                dto.time_to = value['time_to']
                if dto.time_from is None:
                    dto.time_from = 'absent'
                    empColor = 'red'
                else:
                    empColor = timeCompare(dto.time_from)
                
                if empColor == 'green':
                    row_green += 1
                    ws['A'+str(row_green)].value = dto.empid
                    ws['A'+str(row_green)].alignment = Alignment(horizontal='left')
                    ws['A'+str(row_green)].fill = greenFill
                    ws['B'+str(row_green)].value = dto.name
                    ws['B'+str(row_green)].alignment = Alignment(horizontal='left')
                    ws['B'+str(row_green)].fill = greenFill
                    ws['C'+str(row_green)].value = dto.email
                    ws['C'+str(row_green)].alignment = Alignment(horizontal='left')
                    ws['C'+str(row_green)].fill = greenFill
                    ws['D'+str(row_green)].value = dto.time_from
                    ws['D'+str(row_green)].alignment = Alignment(horizontal='left')
                    ws['D'+str(row_green)].fill = greenFill
                    ws['E'+str(row_green)].value = dto.time_to
                    ws['E'+str(row_green)].alignment = Alignment(horizontal='left')
                    ws['E'+str(row_green)].fill = greenFill
                elif empColor == 'yellow':
                    row_yellow += 1
                    ws['G'+str(row_yellow)].value = dto.empid
                    ws['G'+str(row_yellow)].alignment = Alignment(horizontal='left')
                    ws['G'+str(row_yellow)].fill = yellowFill
                    ws['H'+str(row_yellow)].value = dto.name
                    ws['H'+str(row_yellow)].alignment = Alignment(horizontal='left')
                    ws['H'+str(row_yellow)].fill = yellowFill
                    ws['I'+str(row_yellow)].value = dto.email
                    ws['I'+str(row_yellow)].alignment = Alignment(horizontal='left')
                    ws['I'+str(row_yellow)].fill = yellowFill
                    ws['J'+str(row_yellow)].value = dto.time_from
                    ws['J'+str(row_yellow)].alignment = Alignment(horizontal='left')
                    ws['J'+str(row_yellow)].fill = yellowFill
                    ws['K'+str(row_yellow)].value = dto.time_to
                    ws['K'+str(row_yellow)].alignment = Alignment(horizontal='left')
                    ws['K'+str(row_yellow)].fill = yellowFill
                elif empColor =='red':
                    row_red += 1
                    ws['M'+str(row_red)].value = dto.empid
                    ws['M'+str(row_red)].alignment = Alignment(horizontal='left')
                    ws['M'+str(row_red)].fill = redFill
                    ws['N'+str(row_red)].value = dto.name
                    ws['N'+str(row_red)].alignment = Alignment(horizontal='left')
                    ws['N'+str(row_red)].fill = redFill
                    ws['O'+str(row_red)].value = dto.email
                    ws['O'+str(row_red)].alignment = Alignment(horizontal='left')
                    ws['O'+str(row_red)].fill = redFill
                    ws['P'+str(row_red)].value = dto.time_from
                    ws['P'+str(row_red)].alignment = Alignment(horizontal='left')
                    ws['P'+str(row_red)].fill = redFill
                    ws['Q'+str(row_red)].value = dto.time_to
                    ws['Q'+str(row_red)].alignment = Alignment(horizontal='left')
                    ws['Q'+str(row_red)].fill = redFill

        else:
            row_green += 1
            row_yellow +=1
            row_red +=1
            ws.merge_cells('A{0}:E{0}'.format(str(row)))
            ws['A'+str(row)].value = "--- No Records Found ---"
            ws['A'+str(row)].alignment = Alignment(wrapText=True)
            ws['A'+str(row)].alignment = Alignment(horizontal='center')
        wb.save(output)
        return output

class EmployeeDetailsDTO:
    def __init__(self):
        self.id = ''
        self.name = ''
        self.email = ''
        self.time_from = ''
        self.time_to = ''

